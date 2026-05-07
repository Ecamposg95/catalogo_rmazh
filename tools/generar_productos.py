import json
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
CATALOG_DIR = ROOT_DIR / "catalogo"
EXCEL_PATH = CATALOG_DIR / "data" / "productos.xlsx"
JSON_PATH = CATALOG_DIR / "data" / "productos.json"
EXPECTED_COLUMNS = ["nombre", "descripcion", "precio", "categoria", "imagen", "badge", "whatsapp"]
VALID_BADGES = {"", "new", "hot"}


def read_products() -> list[dict]:
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(row=1, column=index).value for index in range(1, len(EXPECTED_COLUMNS) + 1)]

    if headers != EXPECTED_COLUMNS:
        raise ValueError(
            "Encabezados invalidos. Se esperaban: "
            + ", ".join(EXPECTED_COLUMNS)
        )

    products = []
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        product = dict(zip(EXPECTED_COLUMNS, row))
        product = {key: clean_value(value) for key, value in product.items()}

        if not product["nombre"]:
            errors.append(f"Fila {row_number}: nombre vacio")
        if product["precio"] == "":
            errors.append(f"Fila {row_number}: precio vacio")
        else:
            try:
                product["precio"] = float(product["precio"])
            except ValueError:
                errors.append(f"Fila {row_number}: precio no numerico")

        product["badge"] = product["badge"].lower()
        if product["badge"] not in VALID_BADGES:
            errors.append(f"Fila {row_number}: badge debe ser new, hot o vacio")

        product["whatsapp"] = str(product["whatsapp"]).replace("+", "").replace(" ", "")
        products.append(product)

    if errors:
        raise ValueError("Errores encontrados:\n" + "\n".join(errors))

    return products


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No existe {EXCEL_PATH}")

    products = read_products()
    JSON_PATH.write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"Productos exportados: {len(products)}")
    print(f"Archivo generado: {JSON_PATH}")


if __name__ == "__main__":
    main()
